from fastapi import FastAPI
from app.api.v1.endpoints import extraction
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
import io
from openpyxl import Workbook

app = FastAPI()

origins = [
    "http://localhost",
    "http://localhost:4100",
    "https://fp2p-server-hqdfdjerfyfdakbz.eastus-01.azurewebsites.net",
    "https://fp2p-frontend-cwfqadh9egh6fwgd.eastus-01.azurewebsites.net",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(extraction.router, prefix="/api/v1", tags=["Extraction"])

@app.get("/")  # ping básico
def root():
    return {"status": "ok", "service": "fastapi-excel"}

@app.get("/health")  # para health check del App Service
def health():
    return JSONResponse({"status": "healthy"}, status_code=200)

@app.get("/generate-excel")
def generate_excel():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Hello"
    ws["B1"] = "World"

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=generated.xlsx"},
    )